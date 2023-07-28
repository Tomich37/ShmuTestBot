from .database import Database
import configparser
import telebot
import openpyxl
from openpyxl import Workbook
import os
from telebot import types
import re

class Moderation:
    def __init__(self, bot, save_directory):   
        self.database = Database(bot, self.menu_markup)     
        self.bot = bot        
        self.save_directory = save_directory

    # Меню для разных ролей
    @staticmethod
    def admin_markup():
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
        add_moderator_button = types.KeyboardButton(text="Добавить модератора")
        delete_moderator_button = types.KeyboardButton(text="Снять с поста модератора")
        distribution_button= types.KeyboardButton(text="Создать рассылку")
        users_button= types.KeyboardButton(text="Работа с пользователями")
        menu_button= types.KeyboardButton(text="Меню")
        markup.add(distribution_button)
        markup.add(add_moderator_button)
        markup.add(delete_moderator_button)
        markup.add(users_button)
        markup.add(menu_button)
        return markup
    
    @staticmethod
    def admin_user_markup():
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
        add_users_button= types.KeyboardButton(text="Добавить пользователей")
        get_users_button= types.KeyboardButton(text="Выгрузить пользователей")
        users_edit= types.KeyboardButton(text="Редактирование пользователя")
        menu_button= types.KeyboardButton(text="Меню")
        markup.add(add_users_button)
        markup.add(get_users_button)
        markup.add(users_edit)
        markup.add(menu_button)
        return markup

    @staticmethod
    def moder_markup():
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
        distribution_button= types.KeyboardButton(text="Создать рассылку")
        add_data_button= types.KeyboardButton(text="Загрузить материалы")
        users_button= types.KeyboardButton(text="Работа с пользователями")
        menu_button= types.KeyboardButton(text="Меню")
        markup.add(distribution_button)
        markup.add(add_data_button)
        markup.add(users_button)
        markup.add(menu_button)
        return markup

    @staticmethod
    def menu_markup():
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
        distribution_button= types.KeyboardButton(text="Меню")
        markup.add(distribution_button)
        return markup

    # Обработка кнопки "Модерация"
    def moderation_buttonn_klick(self, user_id):
        role = self.database.get_user_role(user_id)
        markup = None          
        if role == "admin":
            markup = self.admin_markup()
            self.bot.send_message(user_id, "Выберите действие:", reply_markup=markup)
        elif role == "moderator":
            markup = self.moder_markup()
            self.bot.send_message(user_id, "Выберите действие:", reply_markup=markup)
        else:      
            markup = self.user_markup()     
            self.bot.send_message(user_id, "У вас недостаточно прав", reply_markup=markup)

    # Обработка кнопки "Добавить модератора"
    def add_moderator_button(self, user_id):
        role = self.database.get_user_role(user_id)
        markup = None         
        if role != "admin":
            markup = self.user_markup()
            self.bot.send_message(user_id, "У вас недостаточно прав", reply_markup=markup)
            if role == "moderator":                
                markup = self.moder_markup()
                self.bot.send_message(user_id, "Выберите действие:", reply_markup=markup)
        else:
            markup = self.admin_markup()
            self.bot.send_message(user_id, text="Команда для назначения модератора: /add_mod \n\n Пример ее использования:\n /add_mod 79998887766", reply_markup=markup)
    
    # Обработка команды /add_mod
    def add_moderator(self, user_id, phone_number):
        role = self.database.get_user_role(user_id)
        markup = None
        if role == "admin":
            user_exists = self.database.user_exists_phone(phone_number)
            if user_exists:
                if self.database.check_admin_role is False:
                    user_info = self.database.user_info(phone_number)
                    # Формируем сообщение с информацией о пользователе
                    user_message = f"Имя: {user_info[2]}\nФамилия: {user_info[3]}\nНомер телефона: {user_info[0]}\nID: {user_info[1]}\nРоль: {user_info[4]}"

                    # Создаем клавиатуру с кнопками "Подтвердить" и "Отмена"
                    keyboard = types.InlineKeyboardMarkup()
                    confirm_add_mod_button = types.InlineKeyboardButton(text="Подтвердить", callback_data=f"confirm_add_mod_{phone_number}")
                    cancel_add_mod_button = types.InlineKeyboardButton(text="Отмена", callback_data="cancel_add_mod")
                    keyboard.add(confirm_add_mod_button, cancel_add_mod_button)

                    # Отправляем сообщение с клавиатурой
                    self.bot.send_message(user_id, user_message, reply_markup=keyboard)
                else:
                    self.bot.send_message(user_id, "Данный пользователь является администратором")
            else:
                self.bot.send_message(user_id, "Пользователь не найден")           
        elif role == "moderator":
            markup = self.moder_markup()
            self.bot.send_message(user_id, "У вас недостаточно прав", reply_markup=markup)
        else:
            markup = self.user_markup()            
            self.bot.send_message(user_id, "У вас недостаточно прав", reply_markup=markup)    

    # Обработка кнопки "Удалить модератора"
    def remove_moderator_button(self, user_id):
        role = self.database.get_user_role(user_id)        
        if role != "admin":
            markup = self.user_markup() 
            self.bot.send_message(user_id, "У вас недостаточно прав", reply_markup=markup)
            if role == "moderator":                
                markup = self.moder_markup()
                self.bot.send_message(user_id, "Выберите действие:", reply_markup=markup)
        else:
            markup = self.admin_markup()
            self.bot.send_message(user_id, text="Команда для снятия модератора: /remove_mod \n\n Пример ее использования:\n /remove_mod 79998887766", reply_markup=markup)

    # Удаление модератора
    def remove_moderator(self, user_id, phone_number):
        role = self.database.get_user_role(user_id)
        markup = None
        if role == "admin":
            user_exists = self.database.user_exists_phone(phone_number)
            if user_exists:
                if self.database.check_admin_role is False:
                    user_info = self.database.user_info(phone_number)
                    # Формируем сообщение с информацией о пользователе
                    user_message = f"Имя: {user_info[2]}\nФамилия: {user_info[3]}\nНомер телефона: {user_info[0]}\nID: {user_info[1]}\nРоль: {user_info[4]}"
                    
                    # Создаем клавиатуру с кнопками "Подтвердить" и "Отмена"
                    keyboard = types.InlineKeyboardMarkup()
                    confirm_remove_mod_button = types.InlineKeyboardButton(text="Подтвердить", callback_data=f"confirm_remove_mod_{phone_number}")
                    cancel_remove_mod_button = types.InlineKeyboardButton(text="Отмена", callback_data="cancel_remove_mod")
                    keyboard.add(confirm_remove_mod_button, cancel_remove_mod_button)

                    # Отправляем сообщение с клавиатурой
                    self.bot.send_message(user_id, user_message, reply_markup=keyboard)
                else:
                    self.bot.send_message(user_id, "Нельзя снять администратора")
            else:
                self.bot.send_message(user_id, "Пользователь не найден")           
        elif role == "moderator":
            markup = self.moder_markup()
            self.bot.send_message(user_id, "У вас недостаточно прав", reply_markup=markup)
        else:
            markup = self.user_markup()            
            self.bot.send_message(user_id, "У вас недостаточно прав", reply_markup=markup)
    
    # Загрузка пользователей
    def add_users(self, message):
        user_id = message.from_user.id
        file_id = message.document.file_id
        file_info = self.bot.get_file(file_id)
        file_name = file_info.file_path.split('/')[-1] 
        downloaded_file = self.bot.download_file(file_info.file_path)
        file_path = os.path.join(self.save_directory, file_name)
        role = self.database.get_user_role(user_id)
        markup = None
        if role == "admin":
            with open(file_path, 'wb') as file:
                file.write(downloaded_file)
            
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            max_column = sheet.max_column

            column_mapping = {
                'phone_number': None,
                'fio': None,
                'region': None,
                'events': None,
                'user_group': None,
                'job': None
            }

            for column in range(1, max_column + 1):
                cell_value = sheet.cell(row=1, column=column).value
                if cell_value in column_mapping:
                    column_mapping[cell_value] = column

            for sheet in workbook.sheetnames:
                current_sheet = workbook[sheet]
                
                for row in range(2, current_sheet.max_row + 1):
                    phone_number = current_sheet.cell(row=row, column=column_mapping['phone_number']).value
                    fio = None
                    region = None
                    user_group = None
                    job = None

                    if phone_number is not None:
                        # Удаление всех букв, символов и пробелов из номера телефона
                        phone_number = re.sub(r'\D', '', str(phone_number))

                    if column_mapping['user_group'] is not None:
                        user_group_cell = current_sheet.cell(row=row, column=column_mapping['user_group'])
                        user_group = user_group_cell.value if user_group_cell.value is not None else None

                    if column_mapping['region'] is not None:
                        region_cell = current_sheet.cell(row=row, column=column_mapping['region'])
                        region = region_cell.value.lower() if region_cell.value is not None else None

                    if column_mapping['fio'] is not None:
                        fio_cell = current_sheet.cell(row=row, column=column_mapping['fio'])
                        fio_value = fio_cell.value
                        if fio_value is not None:
                            words = fio_value.split()
                            fio = ' '.join(words[:2]).lower()
                        else:
                            fio = None
                    else:
                        fio = None

                    if column_mapping['job'] is not None:
                        job_cell = current_sheet.cell(row=row, column=column_mapping['job'])
                        job = job_cell.value.lower() if job_cell.value is not None else None

                    elif phone_number and (phone_number.startswith('8') and len(phone_number) == 11):
                        phone_number = '7' + phone_number[1:]           

                    if phone_number and (phone_number.startswith('7') and len(phone_number) == 11):
                        self.database.insert_or_update_user(phone_number, fio, region, user_group, job)
                    else: 
                        phone_number = None
                        self.database.insert_or_update_user(phone_number, fio, region, user_group, job)


            self.bot.send_message(user_id, "Пользователи добавлены")
            self.database.clear_pending_command(user_id)
        else:
            markup = self.user_markup()            
            self.bot.send_message(user_id, "У вас недостаточно прав", reply_markup=markup)

    def get_users_excel(self, message):
        user_id = message.from_user.id
        role = self.database.get_user_role(user_id)
        markup = None

        if role == "admin":
            users = self.database.get_users_excel()

            workbook = Workbook()
            sheet = workbook.active

            # Запись заголовков столбцов
            sheet.cell(row=1, column=1, value="Phone Number")
            sheet.cell(row=1, column=2, value="FIO")
            sheet.cell(row=1, column=3, value="User Group")
            sheet.cell(row=1, column=4, value="Region")

            # Запись данных пользователей
            for row, user in enumerate(users, start=2):
                phone_number, fio, user_group, region = user
                sheet.cell(row=row, column=1, value=phone_number)
                sheet.cell(row=row, column=2, value=fio)
                sheet.cell(row=row, column=3, value=user_group)
                sheet.cell(row=row, column=4, value=region)

            # Сохранение Excel файла
            current_dir = os.path.dirname(os.path.abspath(__file__))

            # Построение полного пути к файлу фотографии
            temp_dir = os.path.join(current_dir, 'temp')
            exel_path = os.path.join(temp_dir, 'au_uasers.xlsx')
            workbook.save(exel_path)

            markup = self.admin_markup()
            self.bot.send_message(user_id, "Файл с пользователями:", reply_markup=markup)
            self.bot.send_document(user_id, open(exel_path, 'rb'))
        elif role == "moderator":
            users = self.database.get_users_excel()

            workbook = Workbook()
            sheet = workbook.active

            # Запись заголовков столбцов
            sheet.cell(row=1, column=1, value="Phone Number")
            sheet.cell(row=1, column=2, value="FIO")
            sheet.cell(row=1, column=3, value="User Group")
            sheet.cell(row=1, column=4, value="Region")

            # Запись данных пользователей
            for row, user in enumerate(users, start=2):
                phone_number, fio, user_group, region = user

                # Удаление пробелов в конце значения fio
                fio = fio.strip()

                sheet.cell(row=row, column=1, value=phone_number)
                sheet.cell(row=row, column=2, value=fio)
                sheet.cell(row=row, column=3, value=user_group)
                sheet.cell(row=row, column=4, value=region)

            # Сохранение Excel файла
            output_path = "./temp/output.xlsx"
            workbook.save(output_path)

            markup = self.moder_markup()
            self.bot.send_message(user_id, "Файл с пользователями:", reply_markup=markup)
            self.bot.send_document(user_id, open(output_path, 'rb'))
        else:
            markup = self.user_markup()
            self.bot.send_message(user_id, "У вас недостаточно прав", reply_markup=markup)

    def edit_user(self, message):
        user_id = message.from_user.id
        role = self.database.get_user_role(user_id)
        fio = message.text.lower()
        if role != "user":
            user_info = self.database.user_info_fio(fio)
            if user_info is not None:
                # Формируем сообщение с информацией о пользователе
                user_message = f"Информация о пользователе:\n\nФИО: {user_info[2]}\nНомер телефона: {user_info[0]}\nID: {user_info[1]}\nРоль: {user_info[3]}\nРегион: {user_info[4]}\nГруппа: {user_info[5]}"

                phone_number = user_info[0]
                self.database.temp_phone_number(user_id, phone_number)
                
                # Создаем клавиатуру с кнопками "Подтвердить" и "Отмена"
                keyboard = types.InlineKeyboardMarkup()
                edit_user_fio_button = types.InlineKeyboardButton(text="ФИО", callback_data=f"edit_user_fio_{phone_number}")
                edit_user_region_button = types.InlineKeyboardButton(text="Регион", callback_data=f"edit_user_region_{phone_number}")
                edit_user_group_button = types.InlineKeyboardButton(text="Группа", callback_data=f"edit_user_group_{phone_number}")
                cancel_edit_user = types.InlineKeyboardButton(text="Готово", callback_data="edit_user_cancel")
                keyboard.add(edit_user_fio_button, edit_user_region_button, edit_user_group_button, cancel_edit_user)

                # Отправляем сообщение с клавиатурой
                self.bot.send_message(user_id, user_message, reply_markup=keyboard)
            else:
                self.database.clear_pending_command(user_id)
                markup = self.menu_markup()
                self.bot.send_message(user_id, "Пользователь не найден", reply_markup=markup)
        else:
            markup = self.user_markup()            
            self.bot.send_message(user_id, "У вас недостаточно прав", reply_markup=markup)    